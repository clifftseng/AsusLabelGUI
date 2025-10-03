'''
This module handles environment variable loading, initializes API clients, 
and provides helper functions for file processing and Excel reporting.
'''
import os
import json
import base64
import time
import asyncio
import shutil
import re
from pathlib import Path
from collections import Counter

import fitz  # PyMuPDF
from dotenv import load_dotenv
from openai import AzureOpenAI, RateLimitError
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import DocumentAnalysisClient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image

# --- Configuration ---
load_dotenv()

# --- Path Definitions ---
BASE_DIR = Path(__file__).resolve().parent
PROMPT_DIR = BASE_DIR / "prompt"
FORMAT_DIR = BASE_DIR / "format"
REF_DIR = BASE_DIR / "ref"
OUTPUT_DIR = BASE_DIR / "output"
EXCEL_OUTPUT_DIR = OUTPUT_DIR / "excel"

SINGLE_TEMPLATE_PATH = REF_DIR / "single.xlsx"
TOTAL_TEMPLATE_PATH = REF_DIR / "total.xlsx"

# --- Client Cache ---
CACHE = {}

# --- Client Getters ---
def get_azure_openai_client():
    if "aoai_client" in CACHE: return CACHE["aoai_client"]
    endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT")
    api_key = os.environ.get("AZURE_OPENAI_API_KEY")
    deployment = os.environ.get("AZURE_OPENAI_DEPLOYMENT")
    if not all([endpoint, api_key, deployment]): raise ValueError("Azure OpenAI env vars not set.")
    client = AzureOpenAI(api_key=api_key, api_version="2024-02-01", azure_endpoint=endpoint)
    CACHE["aoai_client"] = client
    return client

def get_di_client(log_callback=None):
    if "di_client" in CACHE: return CACHE["di_client"]
    endpoint = os.environ.get("DOCUMENT_INTELLIGENCE_ENDPOINT")
    key = os.environ.get("DOCUMENT_INTELLIGENCE_KEY")
    if not all([endpoint, key]): raise ValueError("Azure DI env vars not set.")
    if log_callback: log_callback("  - 正在建立 DI 用戶端...")
    client = DocumentAnalysisClient(endpoint=endpoint, credential=AzureKeyCredential(key))
    CACHE["di_client"] = client
    if log_callback: log_callback("  - DI 用戶端建立成功")
    return client

def get_azure_openai_deployment():
    deployment = os.environ.get("AZURE_OPENAI_DEPLOYMENT")
    if not deployment: raise ValueError("AZURE_OPENAI_DEPLOYMENT env var not set.")
    return deployment

# --- File/Prompt Readers ---
def read_prompt_file(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as f: return f.read()
    except FileNotFoundError: return None

def get_all_format_keys(log_callback):
    if "all_format_keys" in CACHE: return CACHE["all_format_keys"]
    log_callback("  - 正在掃描所有格式檔以收集目標欄位...")
    all_keys = set()
    try:
        if not FORMAT_DIR.exists():
            log_callback(f"[警告] 格式資料夾 {FORMAT_DIR} 不存在。")
            return set()
        for file_path in FORMAT_DIR.glob('*.json'):
            with open(file_path, 'r', encoding='utf-8') as f: data = json.load(f)
            if isinstance(data.get('hints'), list):
                for hint in data['hints']:
                    if isinstance(hint, dict) and 'field' in hint: all_keys.add(hint['field'])
        log_callback(f"  - 共收集到 {len(all_keys)} 個獨特欄位。")
        CACHE["all_format_keys"] = all_keys
        return all_keys
    except Exception as e:
        log_callback(f"[錯誤] 讀取格式檔時發生錯誤: {e}")
        return set()

# --- API Callers & Helpers ---
async def _query_openai_with_retry(log_callback, **kwargs):
    max_retries, base_delay = 5, 2
    client = get_azure_openai_client()
    for attempt in range(max_retries):
        try:
            loop = asyncio.get_running_loop()
            response = await loop.run_in_executor(None, lambda: client.chat.completions.create(**kwargs))
            log_callback("      - AI 回應接收成功")
            return response
        except RateLimitError as e:
            wait_time = base_delay * (2 ** attempt)
            log_callback(f"    [警告] 遭遇速率限制 (429)。等待 {wait_time} 秒後重試...")
            await asyncio.sleep(wait_time)
        except Exception as e:
            log_callback(f"    [錯誤] AI API 呼叫時發生非預期錯誤: {e}")
            return None
    log_callback(f"    [錯誤] 達到最大重試次數後 API 呼叫失敗")
    return None

async def query_chatgpt_text_api(system_prompt, user_prompt, log_callback, output_dir, file_stem, page_num=None):
    log_callback("  - [ChatGPT] 正在發送純文字請求...")
    
    # Save prompt
    prompt_filename = f"{file_stem}_page_{page_num}_text_prompt.txt" if page_num else f"{file_stem}_text_prompt.txt"
    prompt_filepath = output_dir / prompt_filename
    try:
        with open(prompt_filepath, 'w', encoding='utf-8') as f:
            f.write(f"System Prompt:\n{system_prompt}\n\nUser Prompt:\n{user_prompt}")
        log_callback(f"    - 已儲存 Prompt: {prompt_filename}")
    except Exception as e:
        log_callback(f"    - [錯誤] 儲存 Prompt 失敗: {e}")

    response = await _query_openai_with_retry(log_callback=log_callback, model=get_azure_openai_deployment(), messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}], max_tokens=512, temperature=0.0, response_format={"type": "json_object"})
    
    # Save raw response
    response_filename = f"{file_stem}_page_{page_num}_text_response.json" if page_num else f"{file_stem}_text_response.json"
    response_filepath = output_dir / response_filename
    try:
        if response:
            with open(response_filepath, 'w', encoding='utf-8') as f:
                json.dump(response.model_dump(), f, ensure_ascii=False, indent=4)
            log_callback(f"    - 已儲存原始回應: {response_filename}")
        else:
            log_callback(f"    - 未收到有效回應，未儲存原始回應檔案: {response_filename}")
    except Exception as e:
        log_callback(f"    - [錯誤] 儲存原始回應失敗: {e}")

    if response:
        try:
            return json.loads(response.choices[0].message.content)
        except (json.JSONDecodeError, IndexError, KeyError) as e:
            log_callback(f"    [錯誤] 解析純文字 API 的 JSON 回應時失敗: {e}")
            return None
    return None

async def predict_relevant_pages(pdf_path, log_callback, output_dir):
    log_callback("  - 開始使用 ChatGPT 預測相關頁面...")
    try:
        with fitz.open(pdf_path) as doc: total_pages = len(doc)
        target_fields = get_all_format_keys(log_callback)
        prompt_template = read_prompt_file(PROMPT_DIR / "prompt_page_prediction.txt")
        if not prompt_template: 
            log_callback("  - [錯誤] 找不到頁面預測的 Prompt 檔案。")
            return None
        system_prompt = prompt_template.replace("{TOTAL_PAGES}", str(total_pages)).replace("{TARGET_FIELDS}", "\n".join(sorted(target_fields)))
        user_prompt = "Please provide the JSON object with the most likely pages based on the system prompt."
        
        file_stem = pdf_path.stem
        response_json = await query_chatgpt_text_api(system_prompt, user_prompt, log_callback, output_dir, file_stem, page_num="prediction")
        
        if not isinstance(response_json, dict): 
            log_callback("  - [警告] ChatGPT 沒有回傳有效的 JSON 物件用於頁面預測。")
            return None
        pages = response_json.get("overall_top_pages", [])
        if not pages and isinstance(response_json.get("fields"), dict):
            counter = Counter(p for field_data in response_json["fields"].values() if isinstance(field_data, dict) and isinstance(field_data.get("pages"), list) for p in field_data["pages"] if isinstance(p, int))
            pages = [p for p, _ in counter.most_common(3)]
        cleaned = [p for p in sorted(set(p for p in pages if isinstance(p, int) and 1 <= p <= total_pages))[:3]]
        if cleaned:
            log_callback(f"  - ChatGPT 建議的頁面為: {cleaned}")
            return cleaned
        log_callback("  - [警告] ChatGPT 回應的格式正確但沒有可用頁面。")
        return None
    except Exception as e:
        log_callback(f"[錯誤] 預測相關頁面時發生未預期錯誤: {e}")
        return None

async def analyze_image_with_di(image_path, log_callback):
    try:
        di_client = get_di_client(log_callback)
        log_callback(f"    - [DI] 正在分析圖片: {Path(image_path).name}")
        loop = asyncio.get_running_loop()
        with open(image_path, "rb") as f: image_data = f.read()
        poller = await loop.run_in_executor(None, lambda: di_client.begin_analyze_document("prebuilt-document", document=image_data))
        result = await loop.run_in_executor(None, poller.result)
        log_callback(f"    - [DI] 分析完成")
        return result.to_dict()
    except Exception as e:
        log_callback(f"    - [DI][錯誤] 分析圖片時發生錯誤: {e}")
        return None

def image_file_to_base64(image_path, log_callback):
    try:
        with open(image_path, "rb") as f: return base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        log_callback(f"錯誤：讀取或編碼圖片檔案 '{Path(image_path).name}' 時發生錯誤: {e}")
        return None

async def query_chatgpt_vision_api(system_prompt, user_content, log_callback, output_dir, file_stem, page_num, field=None):
    log_callback("  - 正在發送 Vision API 請求...")

    # Save prompt
    prompt_filename = f"{file_stem}_page_{page_num}_vision_prompt.txt"
    if field: prompt_filename = f"{file_stem}_page_{page_num}_field_{field}_vision_prompt.txt"
    prompt_filepath = output_dir / prompt_filename
    try:
        with open(prompt_filepath, 'w', encoding='utf-8') as f:
            f.write(f"System Prompt:\n{system_prompt}\n\nUser Content:\n{json.dumps(user_content, ensure_ascii=False, indent=4)}")
        log_callback(f"    - 已儲存 Prompt: {prompt_filename}")
    except Exception as e:
        log_callback(f"    - [錯誤] 儲存 Prompt 失敗: {e}")

    response = await _query_openai_with_retry(log_callback=log_callback, model=get_azure_openai_deployment(), messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_content}], max_tokens=4096, temperature=0.1, top_p=0.95, response_format={"type": "json_object"})

    # Save raw response
    response_filename = f"{file_stem}_page_{page_num}_vision_response.json"
    if field: response_filename = f"{file_stem}_page_{page_num}_field_{field}_vision_response.json"
    response_filepath = output_dir / response_filename
    try:
        if response:
            with open(response_filepath, 'w', encoding='utf-8') as f:
                json.dump(response.model_dump(), f, ensure_ascii=False, indent=4)
            log_callback(f"    - 已儲存原始回應: {response_filename}")
        else:
            log_callback(f"    - 未收到有效回應，未儲存原始回應檔案: {response_filename}")
    except Exception as e:
        log_callback(f"    - [錯誤] 儲存原始回應失敗: {e}")

    if response:
        try: return json.loads(response.choices[0].message.content)
        except (json.JSONDecodeError, IndexError, KeyError) as e:
            log_callback(f"    [錯誤] 解析 Vision API 的 JSON 回應時失敗: {e}")
            return None
    return None

async def process_pages_via_screenshot_di_chatgpt(pdf_path, pages_to_process, output_dir, log_callback):
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        log_callback(f"[錯誤] 開啟 PDF 檔案失敗: {e}")
        return []
    output_dir.mkdir(exist_ok=True)
    pdf_stem = pdf_path.stem
    final_chatgpt_results = []
    system_prompt = read_prompt_file(PROMPT_DIR / "prompt_system.txt")
    user_prompt_template = read_prompt_file(PROMPT_DIR / "prompt_user.txt")
    if not all([system_prompt, user_prompt_template]):
        log_callback("[錯誤] 找不到 system/user prompt 檔案，處理終止。")
        doc.close()
        return []
    for page_num in pages_to_process:
        if not (0 <= page_num - 1 < len(doc)):
            log_callback(f"[警告] 頁碼 {page_num} 超出 PDF 範圍，已略過。")
            continue
        try:
            page = doc.load_page(page_num - 1)
            pix = page.get_pixmap(dpi=200)
            output_filename = f"{pdf_stem}_page_{page_num}.png"
            output_filepath = output_dir / output_filename
            pix.save(output_filepath)
            log_callback(f"  - 已儲存截圖: {output_filename}")
            di_result = await analyze_image_with_di(output_filepath, log_callback)
            if not di_result: 
                log_callback(f"  - DI 分析失敗，跳過此頁面。")
                continue
            log_callback(f"  - DI 分析成功。")
            di_content = di_result.get('content', '')
            base64_image = image_file_to_base64(output_filepath, log_callback)
            if not base64_image: 
                log_callback(f"  - [錯誤] Base64 轉換失敗，跳過此頁面。")
                continue
            user_prompt = user_prompt_template.replace("<檔名含副檔名>", pdf_path.name).replace("<整數>", str(len(doc)))
            user_content = [{"type": "text", "text": user_prompt}, {"type": "text", "text": f"\n--- OCRed Text Below ---\n{di_content}"}, {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}}]
            log_callback(f"  - 正在組合提示並呼叫 ChatGPT Vision API...")
            chatgpt_json = await query_chatgpt_vision_api(system_prompt, user_content, log_callback, output_dir, pdf_stem, page_num)
            if chatgpt_json:
                log_callback(f"  - ChatGPT 分析成功。")
                final_chatgpt_results.append({'page': page_num, 'chatgpt_result': chatgpt_json})
            else:
                log_callback(f"  - ChatGPT 分析失敗或沒有回傳結果。")
        except Exception as e:
            log_callback(f"[錯誤] 處理頁面 {page_num} 時失敗: {e}")
    doc.close()
    return final_chatgpt_results

async def process_mode_a_helper(pdf_path, hints, output_dir, log_callback):
    """
    Processes a PDF using hints for cropping, sending both full and cropped images to the AI.
    """
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        log_callback(f"[錯誤] 開啟 PDF 檔案失敗: {e}")
        return []

    output_dir.mkdir(exist_ok=True)
    pdf_stem = pdf_path.stem
    final_chatgpt_results = []

    # Load prompts
    system_prompt = read_prompt_file(PROMPT_DIR / "prompt_system.txt")
    user_prompt_template = read_prompt_file(PROMPT_DIR / "prompt_user.txt")
    if not all([system_prompt, user_prompt_template]):
        log_callback("[錯誤] 找不到 system/user prompt 檔案，處理終止。")
        doc.close()
        return []

    # Group hints by page number to process each page only once
    pages_to_process = {}
    for hint in hints:
        page_num = hint.get('page')
        if not page_num:
            continue
        if page_num not in pages_to_process:
            pages_to_process[page_num] = []
        pages_to_process[page_num].append(hint)

    for page_num, page_hints in pages_to_process.items():
        if not (0 <= page_num - 1 < len(doc)):
            log_callback(f"[警告] 頁碼 {page_num} 超出 PDF 範圍，已略過。")
            continue
        try:
            # --- Step 1: Full Page Screenshot & OCR ---
            page = doc.load_page(page_num - 1)
            pix = page.get_pixmap(dpi=200)
            full_page_filename = f"{pdf_stem}_page_{page_num}_full.png"
            full_page_filepath = output_dir / full_page_filename
            pix.save(full_page_filepath)
            log_callback(f"  - 已儲存全頁截圖: {full_page_filename}")

            di_result = await analyze_image_with_di(full_page_filepath, log_callback)
            if not di_result:
                log_callback(f"  - 全頁 DI 分析失敗，跳過此頁面。")
                continue
            di_content = di_result.get('content', '')
            base64_full_image = image_file_to_base64(full_page_filepath, log_callback)
            if not base64_full_image:
                log_callback(f"  - [錯誤] 全頁 Base64 轉換失敗，跳過此頁面。")
                continue

            # --- Step 2: Process each hint (crop and call AI) on this page ---
            for hint in page_hints:
                field = hint.get('field', 'unknown_field')
                bbox = hint.get('bbox')
                if not bbox or len(bbox) != 4:
                    log_callback(f"  - [警告] 在 hint 中找不到有效的 bbox，跳過欄位 {field} 的裁切。")
                    continue

                try:
                    # --- Crop Image ---
                    with Image.open(full_page_filepath) as img:
                        # bbox is [x, y, width, height]. Convert to (left, upper, right, lower) for Pillow. 
                        x, y, w, h = bbox
                        crop_box = (x, y, x + w, y + h)
                        cropped_img = img.crop(crop_box)
                        cropped_filename = f"{pdf_stem}_page_{page_num}_crop_{field}.png"
                        cropped_filepath = output_dir / cropped_filename
                        cropped_img.save(cropped_filepath)
                        log_callback(f"    - 已儲存裁切圖片: {cropped_filename}")

                    base64_cropped_image = image_file_to_base64(cropped_filepath, log_callback)
                    if not base64_cropped_image:
                        log_callback(f"    - [錯誤] 裁切圖片 Base64 轉換失敗，跳過此欄位。")
                        continue

                    # --- Call ChatGPT Vision API with both images ---
                    log_callback(f"    - 正在為欄位 '{field}' 組合提示並呼叫 ChatGPT Vision API...")
                    user_prompt = user_prompt_template.replace("<檔名含副檔名>", pdf_path.name).replace("<整數>", str(len(doc)))
                    
                    hint_prompt = f"Please analyze the following field: '{field}'. I have provided the full page for context, and a cropped image showing the exact area of interest."

                    user_content = [
                        {"type": "text", "text": user_prompt},
                        {"type": "text", "text": hint_prompt},
                        {"type": "text", "text": f"\n--- OCRed Text Below (from full page) ---\n{di_content}"},
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{base64_full_image}", "detail": "low"}
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{base64_cropped_image}"}
                        }
                    ]

                    chatgpt_json = await query_chatgpt_vision_api(system_prompt, user_content, log_callback, output_dir, pdf_stem, page_num, field)

                    if chatgpt_json:
                        log_callback(f"    - ChatGPT 分析成功 (欄位: {field})。")
                        final_chatgpt_results.append({'page': page_num, 'field': field, 'chatgpt_result': chatgpt_json})
                    else:
                        log_callback(f"    - ChatGPT 分析失敗 (欄位: {field})。")

                except Exception as e:
                    log_callback(f"  - [錯誤] 處理欄位 '{field}' 的 hint 時失敗: {e}")

        except Exception as e:
            log_callback(f"[錯誤] 處理頁面 {page_num} 時失敗: {e}")

    doc.close()
    return final_chatgpt_results
# --- Excel Helper Functions ---
def ensure_template_files_exist(log_callback):
    EXCEL_OUTPUT_DIR.mkdir(exist_ok=True)
    if not (EXCEL_OUTPUT_DIR / "total.xlsx").exists():
        try:
            shutil.copy(TOTAL_TEMPLATE_PATH, EXCEL_OUTPUT_DIR / "total.xlsx")
            log_callback(f"[資訊] 已複製總表範本檔案: total.xlsx")
        except FileNotFoundError:
            log_callback(f"[錯誤] 找不到總表範本檔案: {TOTAL_TEMPLATE_PATH}")
            return False
    return True

def sanitize_for_excel(text):
    if not isinstance(text, str): return text
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)

def get_display_value(data_dict):
    if not isinstance(data_dict, dict): return "無"
    if data_dict.get("value"): return data_dict["value"]
    if data_dict.get("derived_value") is not None: return f"{data_dict['derived_value']} (推論)"
    return "無"

def format_evidence(evidence_list):
    if not evidence_list: return ""
    return "\n".join([f"Page {e.get('page', '?')}: \"{e.get('quote', '')}\"" for e in evidence_list])

def format_conflicts(conflicts_list):
    if not conflicts_list: return ""
    return json.dumps(conflicts_list, ensure_ascii=False, indent=2)

def save_total_excel(all_results, log_callback):
    log_callback("--- 開始儲存總表 Excel ---")

    # Helper to find the category for sorting
    def get_category(result_wrapper):
        if not result_wrapper or not result_wrapper.get('processed_data'):
            return None
        for item in result_wrapper['processed_data']:
            # The category is in the 'file' dictionary of the chatgpt_result
            if item.get('chatgpt_result', {}).get('file', {}).get('category'):
                return item['chatgpt_result']['file']['category']
        return None

    # Sort results to have 'Battery Label Artwork' first
    try:
        all_results.sort(key=lambda res: get_category(res) == 'Battery Label Artwork', reverse=True)
        log_callback("  - 已將 'Battery Label Artwork' 資料置頂排序。")
    except Exception as e:
        log_callback(f"  - [警告] 排序時發生錯誤: {e}")

    try:
        total_output_path = EXCEL_OUTPUT_DIR / "total.xlsx"
        total_wb = load_workbook(total_output_path)
        total_ws = total_wb.active
        for result_wrapper in all_results:
            if not result_wrapper or 'processed_data' not in result_wrapper or not result_wrapper['processed_data']:
                continue
            
            # Merge results from all pages/hints of a single PDF
            merged_data = {}
            for item_result in result_wrapper['processed_data']:
                # The actual data is nested in 'chatgpt_result'
                if 'chatgpt_result' in item_result:
                    merged_data.update(item_result['chatgpt_result'])

            if not merged_data: continue

            # Add file-level info
            if 'file' not in merged_data: merged_data['file'] = {}
            merged_data['file']['name'] = result_wrapper['file_name']
            data = merged_data

            row_data = [
                sanitize_for_excel(data.get('file', {}).get('name', '')),
                sanitize_for_excel(data.get('file', {}).get('category', '')),
                sanitize_for_excel(data.get('model_name', {}).get('value', '')),
                sanitize_for_excel(get_display_value(data.get('nominal_voltage_v', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_batt_capacity_wh', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_energy_wh', {}))),
                sanitize_for_excel(data.get('notes', '')),
                sanitize_for_excel(format_conflicts(data.get('conflicts', [])))
            ]
            total_ws.append(row_data)

        total_wb.save(total_output_path)
        log_callback(f"  - 總表 {total_output_path.name} 已儲存")
        return total_output_path
    except Exception as e:
        log_callback(f"[錯誤] 儲存總表 Excel 檔案時發生錯誤: {e}")
        return None

def apply_highlighting_rules(excel_path, log_callback):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        log_callback(f"讀取 Excel 檔案成功: {os.path.basename(excel_path)}")
    except FileNotFoundError:
        log_callback(f"[錯誤] 找不到要進行標色的 Excel 檔案: {excel_path}")
        return

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    category_col_idx = 2
    columns_to_check_indices = [3, 4, 5, 6, 7, 8]

    artwork_row_indices = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        if len(row) >= category_col_idx and row[category_col_idx - 1] == 'Battery Label Artwork':
            artwork_row_indices.append(row_idx)

    if ws.max_row <= 1:
        log_callback("[資訊] Excel 中沒有資料، 跳過標色 ")
        return

    if len(artwork_row_indices) == 1:
        log_callback("[規則 1] 偵測到單一 'Battery Label Artwork'، 以此為標準 ")
        standard_row_idx = artwork_row_indices[0]
        for col_idx in columns_to_check_indices:
            standard_cell = ws.cell(row=standard_row_idx, column=col_idx)
            standard_value = standard_cell.value
            for row_idx in range(2, ws.max_row + 1):
                if row_idx == standard_row_idx: continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value != standard_value:
                    cell.fill = red_fill

    elif len(artwork_row_indices) > 1:
        log_callback("[規則 2] 偵測到多筆 'Battery Label Artwork'، 進行內部比對 ")
        for col_idx in columns_to_check_indices:
            artwork_values = [ws.cell(row=r_idx, column=col_idx).value for r_idx in artwork_row_indices]
            if len(set(artwork_values)) > 1:
                header_name = ws.cell(row=1, column=col_idx).value
                log_callback(f"  - 欄位 '{header_name}' 在 Artwork 中發現不一致، 全部標紅 ")
                for r_idx in artwork_row_indices:
                    ws.cell(row=r_idx, column=col_idx).fill = red_fill

    else: # No artwork rows
        log_callback("[規則 3] 未偵測到 'Battery Label Artwork'، 採用多數決 ")
        for col_idx in columns_to_check_indices:
            col_values = [ws.cell(row=r_idx, column=col_idx).value for r_idx in range(2, ws.max_row + 1) if ws.cell(row=r_idx, column=col_idx).value is not None]
            if not col_values: continue
            value_counts = Counter(col_values)
            most_common = value_counts.most_common()
            header_name = ws.cell(row=1, column=col_idx).value
            if len(most_common) > 1 and most_common[0][1] == most_common[1][1]:
                log_callback(f"  - 欄位 '{header_name}' 出現平手، 整欄標紅 ")
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=col_idx).fill = red_fill
            else:
                majority_value = most_common[0][0]
                log_callback(f"  - 欄位 '{header_name}' 的多數值為 '{majority_value}' ")
                for r_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    if cell.value != majority_value:
                        cell.fill = red_fill

    wb.save(excel_path)
    log_callback("成功儲存已標色的 Excel 檔案 ")
