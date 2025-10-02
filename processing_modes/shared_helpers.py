

import os
import base64
import json
import re
import sys
import shutil
import time
import threading
from openai import AzureOpenAI, RateLimitError
from dotenv import load_dotenv
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image
from collections import Counter
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import DocumentAnalysisClient
from pathlib import Path

# --- PyInstaller Pathing Helpers ---

def get_base_path():
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if getattr(sys, 'frozen', False):
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        # We need the path to the executable file
        return os.path.dirname(sys.executable)
    else:
        # Not in a PyInstaller bundle
        # os.path.abspath(".") gives the CWD
        return os.path.abspath(".")

def resource_path(relative_path):
    """ Get absolute path to a resource that is bundled with the app """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # Not in a PyInstaller bundle, use the ref folder relative to this script
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "ref"))
        # In dev mode, our structure is different, so we adjust.
        # The original code was BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # and REF_DIR was os.path.join(BASE_DIR, "ref"). So we find the project root.
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        base_path = os.path.join(project_root) # In dev, resource_path will point to project root

    return os.path.join(base_path, relative_path)


# --- Configuration ---
load_dotenv(os.path.join(get_base_path(), '.env')) # Load .env from the exe's directory

# OpenAI
AZURE_OPENAI_ENDPOINT = os.environ.get("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_KEY = os.environ.get("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_DEPLOYMENT_NAME = os.environ.get("AZURE_OPENAI_DEPLOYMENT")
AZURE_OPENAI_API_VERSION = "2024-02-01"
# Document Intelligence
DI_ENDPOINT = os.environ.get("DOCUMENT_INTELLIGENCE_ENDPOINT")
DI_KEY = os.environ.get("DOCUMENT_INTELLIGENCE_KEY")


# --- Dynamic Path Definitions ---
BASE_PATH = get_base_path()

# External directories (relative to the executable)
USER_INPUT_DIR = os.path.join(BASE_PATH, "input")
OUTPUT_DIR = os.path.join(BASE_PATH, "output")
PROMPT_DIR = os.path.join(BASE_PATH, "prompt")
FORMAT_DIR = os.path.join(BASE_PATH, "format")
MODEL_DIR = os.path.join(BASE_PATH, "model")
EXCEL_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "excel")

# Bundled resource files (inside the executable)
# We point to the 'ref' folder which will be bundled.
SINGLE_TEMPLATE_PATH = resource_path(os.path.join("ref", "single.xlsx"))
TOTAL_TEMPLATE_PATH = resource_path(os.path.join("ref", "total.xlsx"))


# --- Model & Client Cache ---
CACHE = {}

def get_device(device_str: str = None):
    import torch
    if device_str:
        return torch.device(device_str)
    if torch.cuda.is_available():
        return torch.device("cuda")
    if hasattr(torch.backends, 'mps') and torch.backends.mps.is_available():
        return torch.device("mps")
    return torch.device("cpu")

def get_owlvit_model(log_callback=None):
    if "owlvit" in CACHE:
        if log_callback: log_callback("  - 從快取中取得 OWL-ViT 模型")
        return CACHE["owlvit"]
    
    from transformers import OwlViTProcessor, OwlViTForObjectDetection
    import torch

    device = get_device()

    try:
        # 優先嘗試從本地資料夾載入
        if log_callback: log_callback(f"  - 正在從本地資料夾 {MODEL_DIR} 載入 OWL-ViT 模型...")
        processor = OwlViTProcessor.from_pretrained(MODEL_DIR)
        model = OwlViTForObjectDetection.from_pretrained(MODEL_DIR).to(device)
        if log_callback: log_callback("  - 本地模型載入成功")

    except OSError:
        # 若本地載入失敗 (例如資料夾不存在)，則從網路下載並儲存
        if log_callback: log_callback(f"[警告] 在 {MODEL_DIR} 中找不到模型。將從網路下載並儲存以供未來使用")
        if log_callback: log_callback("  - 首次下載，請稍候 (可能需要幾分鐘)...")
        
        model_name = "google/owlvit-base-patch32"
        processor = OwlViTProcessor.from_pretrained(model_name)
        model = OwlViTForObjectDetection.from_pretrained(model_name) # 下載
        
        if log_callback: log_callback("  - 模型下載成功، 正在儲存到本地資料夾...")
        if not os.path.exists(MODEL_DIR):
            os.makedirs(MODEL_DIR)
        processor.save_pretrained(MODEL_DIR) # 儲存以供下次使用
        model.save_pretrained(MODEL_DIR)
        if log_callback: log_callback(f"  - 模型已成功儲存至 {MODEL_DIR} ")
        
        model = model.to(device) # 將下載好的模型移至正確裝置
        
    model.eval()
    
    CACHE["owlvit"] = (model, processor)
    if log_callback: log_callback("  - OWL-ViT 模型載入並快取成功")
    return model, processor

def get_azure_openai_client():
    if "aoai_client" in CACHE:
        return CACHE["aoai_client"]
    if not all([AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY, AZURE_OPENAI_DEPLOYMENT_NAME]):
        raise ValueError("Azure OpenAI environment variables are not fully configured.")
    client = AzureOpenAI(api_key=AZURE_OPENAI_API_KEY, api_version=AZURE_OPENAI_API_VERSION, azure_endpoint=AZURE_OPENAI_ENDPOINT)
    CACHE["aoai_client"] = client
    return client

def get_di_client(log_callback=None):
    if "di_client" in CACHE:
        return CACHE["di_client"]
    if not all([DI_ENDPOINT, DI_KEY]):
        raise ValueError("Azure Document Intelligence environment variables are not fully configured.")
    if log_callback: log_callback("  - 正在建立 Document Intelligence 用戶端...")
    client = DocumentAnalysisClient(endpoint=DI_ENDPOINT, credential=AzureKeyCredential(DI_KEY))
    CACHE["di_client"] = client
    if log_callback: log_callback("  - Document Intelligence 用戶端建立成功")
    return client

def preload_models(log_callback=None):
    pass

def ensure_template_files_exist(log_callback):
    if not os.path.exists(EXCEL_OUTPUT_DIR):
        os.makedirs(EXCEL_OUTPUT_DIR)
    target_total_path = os.path.join(EXCEL_OUTPUT_DIR, os.path.basename(TOTAL_TEMPLATE_PATH))
    if not os.path.exists(target_total_path):
        try:
            shutil.copy(TOTAL_TEMPLATE_PATH, target_total_path)
            log_callback(f"[資訊] 已複製總表範本檔案: {os.path.basename(TOTAL_TEMPLATE_PATH)}")
        except FileNotFoundError:
            log_callback(f"[錯誤] 找不到總表範本檔案: {TOTAL_TEMPLATE_PATH}")
            return False
    return True

# --- API Callers & Helpers ---
def sanitize_for_excel(text):
    if not isinstance(text, str):
        return text
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)

def analyze_image_with_di(image_path, log_callback):
    """Analyzes an image file with Azure Document Intelligence."""
    try:
        di_client = get_di_client(log_callback)
        log_callback(f"    - [DI] 正在分析圖片: {os.path.basename(image_path)}")
        with open(image_path, "rb") as f:
            poller = di_client.begin_analyze_document("prebuilt-document", document=f)
        result = poller.result()
        log_callback(f"    - [DI] 分析完成")
        return result.to_dict()
    except Exception as e:
        log_callback(f"    - [DI][錯誤] 分析圖片時發生錯誤: {e}")
        return None

def _query_openai_with_retry(log_callback, **kwargs):
    """Internal helper to query OpenAI with exponential backoff."""
    max_retries = 5
    base_delay = 2  # seconds

    for attempt in range(max_retries):
        try:
            client = get_azure_openai_client()
            response = client.chat.completions.create(**kwargs)
            log_callback("      - AI 回應接收成功")
            return response
        except RateLimitError as e:
            wait_time = base_delay * (2 ** attempt)
            log_callback(f"    [警告] 遭遇速率限制 (429)。正在等待 {wait_time} 秒後重試... (第 {attempt + 1}/{max_retries} 次)")
            time.sleep(wait_time)
        except Exception as e:
            log_callback(f"    [錯誤] AI API 呼叫時發生非預期錯誤: {e}")
            return None # For non-retryable errors
    
    log_callback(f"    [錯誤] 達到最大重試次數 ({max_retries}) 後، API 呼叫仍然失敗")
    return None

def query_chatgpt_vision_api(system_prompt, user_content, log_callback):
    log_callback("  - 正在發送 Vision API 請求...")
    response = _query_openai_with_retry(
        log_callback=log_callback,
        model=AZURE_OPENAI_DEPLOYMENT_NAME,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ],
        max_tokens=4096, 
        temperature=0.1, 
        top_p=0.95, 
        response_format={"type": "json_object"}
    )
    if response:
        try:
            return json.loads(response.choices[0].message.content)
        except (json.JSONDecodeError, IndexError, KeyError) as e:
            log_callback(f"    [錯誤] 解析 Vision API 的 JSON 回應時失敗: {e}")
            return None
    return None

def query_chatgpt_text_api(system_prompt, user_prompt, log_callback):
    log_callback("  - [ChatGPT] 正在發送純文字請求以預測頁面...")
    response = _query_openai_with_retry(
        log_callback=log_callback,
        model=AZURE_OPENAI_DEPLOYMENT_NAME,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        max_tokens=512,
        temperature=0.0,
        response_format={"type": "json_object"}
    )
    if response:
        try:
            response_str = response.choices[0].message.content
            # 直接載入 JSON 字串，因為 response_format={"type": "json_object"} 應確保其為有效的 JSON
            return json.loads(response_str)
        except (json.JSONDecodeError, IndexError, KeyError) as e:
            log_callback(f"    [錯誤] 解析純文字 API 的 JSON 回應時失敗: {e}")
            log_callback(f"      - 收到的原始字串: '{response.choices[0].message.content}'")
            return None
    return None

def get_all_format_keys(log_callback):
    if "all_format_keys" in CACHE:
        return CACHE["all_format_keys"]

    log_callback("  - 正在掃描所有格式檔以收集目標欄位...")
    all_keys = set()
    try:
        json_files = [f for f in os.listdir(FORMAT_DIR) if f.lower().endswith('.json')]
        for file_name in json_files:
            file_path = os.path.join(FORMAT_DIR, file_name)
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data.get('hints'), list):
                for hint in data['hints']:
                    if isinstance(hint, dict) and 'field' in hint:
                        all_keys.add(hint['field'])
        log_callback(f"  - 共收集到 {len(all_keys)} 個獨特欄位。".encode('cp950', errors='replace').decode('cp950'))
        CACHE["all_format_keys"] = all_keys
        return all_keys
    except Exception as e:
        safe_error_msg = str(e).encode('cp950', errors='replace').decode('cp950')
        log_callback(f"[錯誤] 讀取格式檔以收集鍵時發生錯誤: {safe_error_msg}")
        return set()


def predict_relevant_pages(pdf_path, log_callback):
    """Uses ChatGPT to predict the most relevant pages in a document."""
    log_callback("  - 開始使用 ChatGPT 預測相關頁面...")
    try:
        import fitz
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        doc.close()

        #（可留可去）收集欄位，僅作為日後擴充；目前的 prompt 不使用 {TARGET_FIELDS}
        target_fields = get_all_format_keys(log_callback)
        if not target_fields:
            log_callback("  - [警告] 找不到任何目標欄位、無法進行頁面預測")
            return None

        prompt_template = read_prompt_file(os.path.join(PROMPT_DIR, "prompt_page_prediction.txt"))
        if not prompt_template:
            log_callback("  - [錯誤] 找不到頁面預測的 Prompt 檔案")
            return None

        # ⚠ 只替換我們真的需要的占位符，避免 format 吃掉 JSON 大括號
        system_prompt = prompt_template.replace("{TOTAL_PAGES}", str(total_pages))
        # 若未來你的模板真的加入 {TARGET_FIELDS}，這行會安全替換；現在不影響
        system_prompt = system_prompt.replace("{TARGET_FIELDS}", "\n".join(sorted(target_fields)))

        user_prompt = "Please provide the JSON object with the most likely pages based on the system prompt."

        response_json = query_chatgpt_text_api(system_prompt, user_prompt, log_callback)

        log_callback(f"--- DEBUG: response_json 的型別: {type(response_json)} ---")
        log_callback(f"--- DEBUG: response_json 的內容 (repr): {repr(response_json).encode('cp950', errors='replace').decode('cp950')} ---")

        if not isinstance(response_json, dict):
            log_callback("  - [警告] ChatGPT 沒有回傳有效的 JSON 物件")
            return None

        # 1) 先拿「官方」匯總欄位
        pages = []
        if isinstance(response_json.get("overall_top_pages"), list):
            pages = response_json["overall_top_pages"]

        # 2) 若沒有 overall_top_pages，就從 fields.*.pages 做投票聚合
        if (not pages) and isinstance(response_json.get("fields"), dict):
            from collections import Counter
            counter = Counter()
            for field_data in response_json["fields"].values():
                if isinstance(field_data, dict) and isinstance(field_data.get("pages"), list):
                    for p in field_data["pages"]:
                        if isinstance(p, int):
                            counter[p] += 1
            pages = [p for p, _ in counter.most_common(3)]

        # 3) 最後才回退到舊鍵名（向後相容）
        if (not pages) and isinstance(response_json.get("most_likely_pages"), list):
            pages = response_json["most_likely_pages"]

        # 合法性檢查：1-based 且不能超界；去重後保序取前 3
        def valid(p): return isinstance(p, int) and (1 <= p <= total_pages)
        seen, cleaned = set(), []
        for p in pages:
            if valid(p) and p not in seen:
                seen.add(p)
                cleaned.append(p)
            if len(cleaned) >= 3:
                break

        if cleaned:
            log_callback(f"  - ChatGPT 建議的頁面為: {cleaned}")
            return cleaned

        log_callback("  - [警告] ChatGPT 回應的格式正確但沒有可用頁面")
        return None

    except Exception as e:
        log_callback(f"[錯誤] 預測相關頁面時發生未預期錯誤: {e}")
        return None


def read_prompt_file(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return None

def image_file_to_base64(image_path, log_callback):
    try:
        with open(image_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        log_callback(f"錯誤：讀取或編碼圖片檔案 '{os.path.basename(image_path)}' 時發生錯誤: {e}")
        return None

def pdf_to_base64_images(pdf_path, log_callback, sub_progress_callback=None, pages_to_process=None):
    """Converts specified pages of a PDF to a list of base64 encoded images."""
    base64_images = []
    try:
        doc = fitz.open(pdf_path)
        total_pages_in_doc = len(doc)

        if pages_to_process:
            target_pages = [p - 1 for p in pages_to_process if 0 < p <= total_pages_in_doc]
            if not target_pages:
                log_callback(f"  - [警告] 指定的頁面 {pages_to_process} 在 PDF 中均無效 (總頁數: {total_pages_in_doc}) ")
                return []
            log_callback(f"  - 將處理指定的 {len(target_pages)} 頁: {[p + 1 for p in target_pages]}")
        else:
            target_pages = range(total_pages_in_doc)
            log_callback(f"  - 將處理所有 {total_pages_in_doc} 頁 ")

        total_pages_to_convert = len(target_pages)
        for i, page_num in enumerate(target_pages):
            try:
                page = doc.load_page(page_num)
                pix = page.get_pixmap(dpi=150)
                img_bytes = pix.tobytes("png")
                base64_images.append(base64.b64encode(img_bytes).decode('utf-8'))
                if sub_progress_callback:
                    sub_progress_callback(i + 1, total_pages_to_convert)
            except Exception as e:
                log_callback(f"  - [錯誤] 處理頁面 {page_num + 1} 時失敗: {e}")
        
        doc.close()
        return base64_images

    except Exception as e:
        log_callback(f"[錯誤] 開啟或讀取 PDF 檔案 '{os.path.basename(pdf_path)}' 時發生錯誤: {e}")
        return None


# --- Excel Helper Functions ---
def get_display_value(data_dict):
    if not isinstance(data_dict, dict):
        return "無"
    if data_dict.get("value"):
        return data_dict["value"]
    if data_dict.get("derived_value") is not None:
        return f"{data_dict['derived_value']} (推論)"
    return "無"

def format_evidence(evidence_list):
    if not evidence_list:
        return ""
    return "\n".join([
        f"Page {e.get('page', '?')} (loc: {e.get('loc', 'N/A')}): \"{e.get('quote', '')}\""
        for e in evidence_list
    ])

def format_conflicts(conflicts_list):
    if not conflicts_list:
        return ""
    return json.dumps(conflicts_list, ensure_ascii=False, indent=2)

def save_single_excel(processed_data_wrapper, output_folder, log_callback):
    if not processed_data_wrapper or 'processed_data' not in processed_data_wrapper or not processed_data_wrapper['processed_data']:
        log_callback(f"[警告] 無法為檔案 {processed_data_wrapper.get('file_name', '未知')} 儲存 Excel، 因為沒有有效的處理資料 ")
        return

    original_filename = processed_data_wrapper['file_name']
    all_json_results = processed_data_wrapper['processed_data']
    file_name_without_ext = os.path.splitext(original_filename)[0]

    merged_data = {}
    for d in all_json_results:
        merged_data.update(d)
    
    if 'file' not in merged_data:
        merged_data['file'] = {}
    merged_data['file']['name'] = original_filename

    data = merged_data

    try:
        single_output_path = os.path.join(output_folder, f"single_{file_name_without_ext}.xlsx")
        if not os.path.exists(single_output_path):
             shutil.copy(SINGLE_TEMPLATE_PATH, single_output_path)

        single_wb = load_workbook(single_output_path)
        ws = single_wb.active

        ws['B2'] = sanitize_for_excel(data.get('file', {}).get('name', ''))
        ws['B3'] = sanitize_for_excel(data.get('file', {}).get('category', ''))
        ws['B4'] = sanitize_for_excel(data.get('model_name', {}).get('value', '無'))
        ws['C4'] = sanitize_for_excel(format_evidence(data.get('model_name', {}).get('evidence', [])))

        fields_map = {
            'nominal_voltage_v': ('B5', 'C5'), 'typ_batt_capacity_wh': ('B6', 'C6'),
            'typ_capacity_mah': ('B7', 'C7'), 'rated_capacity_mah': ('B8', 'C8'),
            'rated_energy_wh': ('B9', 'C9'),
        }
        for key, (val_cell, evi_cell) in fields_map.items():
            field_data = data.get(key, {})
            ws[val_cell] = sanitize_for_excel(get_display_value(field_data))
            ws[evi_cell] = sanitize_for_excel(format_evidence(field_data.get('evidence', [])))
        
        ws['B13'] = sanitize_for_excel(data.get('notes', ''))
        ws['B15'] = sanitize_for_excel(format_conflicts(data.get('conflicts', [])))
        
        single_wb.save(single_output_path)
        log_callback(f"  - 已儲存單一 Excel 檔案: {os.path.basename(single_output_path)}")

    except Exception as e:
        log_callback(f"[錯誤] 儲存單一 Excel 檔案時發生錯誤 ({original_filename}): {e}")

def save_total_excel(all_results, output_folder, log_callback):
    log_callback("--- 開始儲存總表 Excel ---")
    try:
        total_output_path = os.path.join(output_folder, "total.xlsx")
        if not os.path.exists(total_output_path):
            shutil.copy(TOTAL_TEMPLATE_PATH, total_output_path)

        total_wb = load_workbook(total_output_path)
        total_ws = total_wb.active

        for result_wrapper in all_results:
            if not result_wrapper or 'processed_data' not in result_wrapper or not result_wrapper['processed_data']:
                continue

            merged_data = {}
            for d in result_wrapper['processed_data']:
                merged_data.update(d)
            
            if 'file' not in merged_data:
                merged_data['file'] = {}
            merged_data['file']['name'] = result_wrapper['file_name']
            data = merged_data

            row_data = [
                sanitize_for_excel(data.get('file', {}).get('name', '')),
                sanitize_for_excel(data.get('file', {}).get('category', '')),
                sanitize_for_excel(data.get('model_name', {}).get('value', '')),
                sanitize_for_excel(get_display_value(data.get('nominal_voltage_v', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_batt_capacity_wh', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_energy_wh', {}))),
                sanitize_for_excel(data.get('notes', '')),
                sanitize_for_excel(format_conflicts(data.get('conflicts', [])))
            ]
            total_ws.append(row_data)

        total_wb.save(total_output_path)
        log_callback("  - 總表 total.xlsx 已儲存 ")
        return total_output_path # Return path for further processing

    except Exception as e:
        log_callback(f"[錯誤] 儲存總表 Excel 檔案時發生錯誤: {e}")
        return None


def apply_highlighting_rules(excel_path, log_callback):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        log_callback(f"讀取 Excel 檔案成功: {os.path.basename(excel_path)}")
    except FileNotFoundError:
        log_callback(f"[錯誤] 找不到要進行標色的 Excel 檔案: {excel_path}")
        return

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    category_col_idx = 2
    columns_to_check_indices = [3, 4, 5, 6, 7, 8]

    artwork_row_indices = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        if len(row) >= category_col_idx and row[category_col_idx - 1] == 'Battery Label Artwork':
            artwork_row_indices.append(row_idx)

    if ws.max_row <= 1:
        log_callback("[資訊] Excel 中沒有資料، 跳過標色 ")
        return

    if len(artwork_row_indices) == 1:
        log_callback("[規則 1] 偵測到單一 'Battery Label Artwork'، 以此為標準 ")
        standard_row_idx = artwork_row_indices[0]
        for col_idx in columns_to_check_indices:
            standard_cell = ws.cell(row=standard_row_idx, column=col_idx)
            standard_value = standard_cell.value
            for row_idx in range(2, ws.max_row + 1):
                if row_idx == standard_row_idx: continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value != standard_value:
                    cell.fill = red_fill

    elif len(artwork_row_indices) > 1:
        log_callback("[規則 2] 偵測到多筆 'Battery Label Artwork'، 進行內部比對 ")
        for col_idx in columns_to_check_indices:
            artwork_values = [ws.cell(row=r_idx, column=col_idx).value for r_idx in artwork_row_indices]
            if len(set(artwork_values)) > 1:
                header_name = ws.cell(row=1, column=col_idx).value
                log_callback(f"  - 欄位 '{header_name}' 在 Artwork 中發現不一致، 全部標紅 ")
                for r_idx in artwork_row_indices:
                    ws.cell(row=r_idx, column=col_idx).fill = red_fill

    else: # No artwork rows
        log_callback("[規則 3] 未偵測到 'Battery Label Artwork'، 採用多數決 ")
        for col_idx in columns_to_check_indices:
            col_values = [ws.cell(row=r_idx, column=col_idx).value for r_idx in range(2, ws.max_row + 1) if ws.cell(row=r_idx, column=col_idx).value is not None]
            if not col_values: continue
            value_counts = Counter(col_values)
            most_common = value_counts.most_common()
            header_name = ws.cell(row=1, column=col_idx).value
            if len(most_common) > 1 and most_common[0][1] == most_common[1][1]:
                log_callback(f"  - 欄位 '{header_name}' 出現平手، 整欄標紅 ")
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=col_idx).fill = red_fill
            else:
                majority_value = most_common[0][0]
                log_callback(f"  - 欄位 '{header_name}' 的多數值為 '{majority_value}' ")
                for r_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    if cell.value != majority_value:
                        cell.fill = red_fill

    wb.save(excel_path)
    log_callback("成功儲存已標色的 Excel 檔案 ")


# --- NEW SHARED FUNCTION ---
def process_pages_via_screenshot_di_chatgpt(pdf_path, pages_to_process, output_dir, log_callback):
    """
    A shared helper function that takes a list of pages, screenshots them,
    runs DI, and then runs ChatGPT Vision analysis.
    """
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        log_callback(f"[錯誤] 開啟 PDF 檔案失敗: {e}")
        return []

    os.makedirs(output_dir, exist_ok=True)
    pdf_stem = Path(pdf_path).stem
    
    final_chatgpt_results = []

    # 讀取通用的 prompt 檔案
    system_prompt = read_prompt_file(os.path.join(PROMPT_DIR, "prompt_system.txt"))
    user_prompt_template = read_prompt_file(os.path.join(PROMPT_DIR, "prompt_user.txt"))

    if not all([system_prompt, user_prompt_template]):
        log_callback("[錯誤] 找不到 prompt_system.txt 或 prompt_user.txt، 處理終止 ")
        doc.close()
        return []

    for page_num in pages_to_process:
        page_index = page_num - 1
        if 0 <= page_index < len(doc):
            try:
                # --- 步驟 a: 截圖 ---
                page = doc.load_page(page_index)
                pix = page.get_pixmap(dpi=200)
                output_filename = f"{pdf_stem}_page_{page_num}.png"
                output_filepath = os.path.join(output_dir, output_filename)
                pix.save(output_filepath)
                log_callback(f"  - 已儲存截圖: {output_filename}")

                # --- 步驟 b: DI 分析 ---
                log_callback(f"  - 正在將 {output_filename} 送往 Document Intelligence...")
                di_result = analyze_image_with_di(output_filepath, log_callback)
                if not di_result:
                    log_callback(f"  - DI 分析失敗或沒有回傳結果، 跳過此頁面 ")
                    continue
                
                log_callback(f"  - DI 分析成功 ")
                di_content = di_result.get('content', '')

                # --- 步驟 c: 呼叫 ChatGPT Vision API ---
                log_callback(f"  - 正在組合提示並呼叫 ChatGPT Vision API...")
                
                base64_image = image_file_to_base64(output_filepath, log_callback)
                if not base64_image:
                    log_callback(f"  - [錯誤] 無法將圖片轉為 Base64، 跳過此頁面 ")
                    continue

                user_content = [
                    {"type": "text", "text": user_prompt_template},
                    {"type": "text", "text": f"\n--- OCRed Text Below ---\n{di_content}"},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}}
                ]

                chatgpt_json = query_chatgpt_vision_api(system_prompt, user_content, log_callback)

                if chatgpt_json:
                    log_callback(f"  - ChatGPT 分析成功 ")
                    final_chatgpt_results.append({'page': page_num, 'chatgpt_result': chatgpt_json})
                else:
                    log_callback(f"  - ChatGPT 分析失敗或沒有回傳結果 ")

            except Exception as e:
                log_callback(f"[錯誤] 處理頁面 {page_num} 時失敗: {e}")
        else:
            log_callback(f"[警告] 頁碼 {page_num} 超出 PDF 範圍 (總頁數: {len(doc)})، 已略過 ")

    doc.close()
    return final_chatgpt_results
