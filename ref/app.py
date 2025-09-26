import os
import base64
import json
import shutil
import re
import sys
from openai import AzureOpenAI
from dotenv import load_dotenv
import fitz  # PyMuPDF
from openpyxl import load_workbook
from PIL import Image

# --- Configuration ---
# Load environment variables from .env file
load_dotenv()

# Get Azure OpenAI credentials from environment variables
try:
    AZURE_OPENAI_ENDPOINT = os.environ["AZURE_OPENAI_ENDPOINT"]
    AZURE_OPENAI_API_KEY = os.environ["AZURE_OPENAI_API_KEY"]
    AZURE_OPENAI_DEPLOYMENT_NAME = os.environ["AZURE_OPENAI_DEPLOYMENT"]
    AZURE_OPENAI_API_VERSION = "2024-02-01" # Use a recent API version that supports vision
except KeyError as e:
    print(f"錯誤：請確認您的 .env 檔案中已設定好 {e} 這個環境變數。")
    exit()

# Define directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USER_INPUT_DIR = os.path.join(BASE_DIR, "user_input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
PROMPT_DIR = os.path.join(BASE_DIR, "prompt")
EXCEL_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "excel")
SINGLE_TEMPLATE_PATH = os.path.join(BASE_DIR, "single.xlsx")
TOTAL_TEMPLATE_PATH = os.path.join(BASE_DIR, "total.xlsx")


# --- Helper Functions ---
def sanitize_for_excel(text):
    """Removes illegal characters for XML/Excel from a string."""
    if not isinstance(text, str):
        return text
    # XML 1.0 spec forbids characters 0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)

def encode_image_to_base64(image_bytes):
    """Encodes image bytes to a base64 string."""
    return base64.b64encode(image_bytes).decode("utf-8")

def pdf_to_base64_images(pdf_path):
    """Converts each page of a PDF to a list of base64 encoded image strings."""
    images = []
    try:
        doc = fitz.open(pdf_path)
        for page_num, page in enumerate(doc):
            pix = page.get_pixmap(dpi=150)
            img_bytes = pix.tobytes("png")
            images.append(encode_image_to_base64(img_bytes))
            print(f"  - 已轉換第 {page_num + 1}/{len(doc)} 頁...")
        doc.close()
    except Exception as e:
        print(f"處理 PDF '{os.path.basename(pdf_path)}' 時發生錯誤: {e}")
        return None
    return images

def image_file_to_base64(image_path):
    """Encodes an image file to a base64 string."""
    try:
        with open(image_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        print(f"錯誤：讀取或編碼圖片檔案 '{os.path.basename(image_path)}' 時發生錯誤: {e}")
        return None

def read_prompt_file(file_path):
    """Reads content from a prompt file."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        print(f"錯誤：找不到提示檔案 {file_path}")
        return ""

# --- Excel Helper Functions ---
def get_display_value(data_dict):
    """Gets the value to display in Excel, prioritizing value, then derived_value."""
    if not isinstance(data_dict, dict):
        return "無"
    if data_dict.get("value"):
        return data_dict["value"]
    if data_dict.get("derived_value") is not None:
        return f"{data_dict['derived_value']} (推論)"
    return "無"

def format_evidence(evidence_list):
    """Formats the evidence list into a readable string."""
    if not evidence_list:
        return ""
    return "\n".join([
        f"Page {e.get('page', '?')} (loc: {e.get('loc', 'N/A')}): \"{e.get('quote', '')}\""
        for e in evidence_list
    ])

def format_conflicts(conflicts_list):
    """Formats the conflicts list into a readable string."""
    if not conflicts_list:
        return ""
    return json.dumps(conflicts_list, ensure_ascii=False, indent=2)

# --- Main Logic ---
def method_purellm():
    """Main function to process PDFs, query Azure OpenAI, and generate Excel reports incrementally."""

    # 1. Initial Setup
    print("--- 清空 output 目錄 ---")
    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR)
    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)
    print("output 目錄已清空並重建。")

    if not os.path.exists(SINGLE_TEMPLATE_PATH) or not os.path.exists(TOTAL_TEMPLATE_PATH):
        print(f"錯誤: 找不到範本檔案 single.xlsx 或 total.xlsx。請確認檔案位於 {BASE_DIR}")
        return
    
    shutil.copy(TOTAL_TEMPLATE_PATH, os.path.join(EXCEL_OUTPUT_DIR, "total.xlsx"))

    print("\n--- 開始增量處理 PDF 檔案 ---")

    client = AzureOpenAI(
        api_key=AZURE_OPENAI_API_KEY,
        api_version=AZURE_OPENAI_API_VERSION,
        azure_endpoint=AZURE_OPENAI_ENDPOINT
    )
    system_prompt = read_prompt_file(os.path.join(PROMPT_DIR, "prompt_system.txt"))
    user_prompt_template = read_prompt_file(os.path.join(PROMPT_DIR, "prompt_user.txt"))

    if not system_prompt or not user_prompt_template:
        print("錯誤：無法讀取必要的提示檔案，程式終止。")
        return

    pdf_files = [f for f in os.listdir(USER_INPUT_DIR) if f.lower().endswith(".pdf")]
    if not pdf_files:
        print(f"在 {USER_INPUT_DIR} 中找不到任何 PDF 檔案。")
        return

    # 2. Main Incremental Loop
    for filename in pdf_files:
        print(f"\n--- 正在處理檔案: {filename} ---")

        base64_images = pdf_to_base64_images(pdf_path=os.path.join(USER_INPUT_DIR, filename))
        if not base64_images:
            continue

        page_count = len(base64_images)
        current_user_prompt = user_prompt_template.replace("<檔名含副檔名>", filename).replace("<整數>", str(page_count))
        print(f"  - 動態產生使用者提示，檔名: {filename}, 頁數: {page_count}")

        all_json_results = []
        for i in range(0, len(base64_images), 50):
            batch_images = base64_images[i : i + 50]
            print(f"  - 正在處理批次 (頁面 {i+1} 到 {min(i + 50, page_count)})... ")
            
            user_content = [{"type": "text", "text": current_user_prompt}]
            user_content.extend([{"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img}"}} for img in batch_images])

            try:
                print(f"  - 正在向 Azure OpenAI 發送請求 ({len(batch_images)} 張圖片)... ")
                response = client.chat.completions.create(
                    model=AZURE_OPENAI_DEPLOYMENT_NAME,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_content}
                    ],
                    max_tokens=4096, temperature=0.1, top_p=0.95, response_format={"type": "json_object"}
                )
                all_json_results.append(json.loads(response.choices[0].message.content))
                print("  - 成功收到回應。")
            except Exception as e:
                print(f"  - 呼叫 Azure OpenAI API 時發生錯誤: {e}")

        if not all_json_results:
            print(f"檔案 {filename} 沒有產生任何有效的 JSON 結果，跳過後續儲存步驟。")
            continue

        # All subsequent file operations are grouped here for robustness
        try:
            # 1. Prepare and save the JSON data
            merged_json = {k: v for d in all_json_results for k, v in d.items()}
            if 'file' not in merged_json: merged_json['file'] = {}
            merged_json['file']['name'] = filename
            print(f"  - 已使用實際檔案名稱 '{filename}' 覆寫 file.name。")
            
            json_output_filename = os.path.splitext(filename)[0] + ".json"
            with open(os.path.join(OUTPUT_DIR, json_output_filename), "w", encoding="utf-8") as f:
                json.dump(merged_json, f, ensure_ascii=False, indent=4)
            print(f"  - 已儲存 JSON 檔案: {json_output_filename}")

            data = merged_json

            # 2. Generate Single Excel
            single_wb = load_workbook(SINGLE_TEMPLATE_PATH)
            ws = single_wb.active
            ws['B2'] = sanitize_for_excel(data.get('file', {}).get('name', ''))
            ws['B3'] = sanitize_for_excel(data.get('file', {}).get('category', ''))
            
            ws['B4'] = sanitize_for_excel(data.get('model_name', {}).get('value', '無'))
            ws['C4'] = sanitize_for_excel(format_evidence(data.get('model_name', {}).get('evidence', [])))

            fields_map = {
                'nominal_voltage_v': ('B5', 'C5'),
                'typ_batt_capacity_wh': ('B6', 'C6'), 'typ_capacity_mah': ('B7', 'C7'),
                'rated_capacity_mah': ('B8', 'C8'), 'rated_energy_wh': ('B9', 'C9'),
            }
            for key, (val_cell, evi_cell) in fields_map.items():
                field_data = data.get(key, {})
                ws[val_cell] = sanitize_for_excel(get_display_value(field_data))
                ws[evi_cell] = sanitize_for_excel(format_evidence(field_data.get('evidence', [])))
            
            ws['B13'] = sanitize_for_excel(data.get('notes', ''))
            ws['B15'] = sanitize_for_excel(format_conflicts(data.get('conflicts', [])))
            
            excel_filename = os.path.splitext(filename)[0] + ".xlsx"
            single_output_path = os.path.join(EXCEL_OUTPUT_DIR, excel_filename)
            single_wb.save(single_output_path)
            print(f"  - 已儲存單一 Excel 檔案: {excel_filename}")

            # 3. Append to Total Excel and Save
            total_output_path = os.path.join(EXCEL_OUTPUT_DIR, "total.xlsx")
            total_wb = load_workbook(total_output_path)
            total_ws = total_wb.active
            row_data = [
                sanitize_for_excel(data.get('file', {}).get('name', '')), 
                sanitize_for_excel(data.get('file', {}).get('category', '')),
                sanitize_for_excel(data.get('model_name', {}).get('value', '')),
                sanitize_for_excel(get_display_value(data.get('nominal_voltage_v', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_batt_capacity_wh', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_energy_wh', {}))),
                sanitize_for_excel(data.get('notes', '')),
                sanitize_for_excel(format_conflicts(data.get('conflicts', [])))
            ]
            total_ws.append(row_data)
            total_wb.save(total_output_path)
            print(f"  - 已更新並儲存 total.xlsx")

        except Exception as e:
            error_message = f"  - 處理檔案 {filename} 的後續儲存（JSON/Excel）時發生嚴重錯誤: {e}"
            safe_error_message = error_message.encode('utf-8', 'replace').decode(sys.stdout.encoding, 'replace')
            print(safe_error_message)
            continue

    print("\n--- 所有檔案處理完畢 ---")

# --- Image Processing Logic (New) ---

def method_llm_with_label():
    """
    主函數：根據預定義的 JSON 格式處理 PDF 檔案。
    流程包括：
    1. 讀取 format 目錄下的 JSON 設定檔，建立格式映射。
    2. 初始化 Azure OpenAI 客戶端和系統提示。
    3. 遍歷 user_input 目錄下的每個 PDF 檔案。
    4. 對每個 PDF 檔案，嘗試匹配其名稱與預定義的格式。
    5. 如果匹配成功，則根據 JSON 設定對 PDF 進行圖像處理（縮放、裁剪）。
    6. 將處理後的圖像發送至 Azure OpenAI 進行標籤提取。
    7. 將 Azure OpenAI 的回應儲存為 JSON 檔案。
    每個 PDF 檔案會獨立完成上述所有步驟，然後再處理下一個檔案。
    """
    print("--- 開始根據 format JSON 處理圖片並呼叫 Azure OpenAI ---")

    # 確保 output 目錄存在，如果不存在則建立
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"已建立 output 目錄: {OUTPUT_DIR}")

    # 1. 載入 format 目錄下的所有 JSON 設定檔
    # 建立一個不區分大小寫的格式名稱到檔案路徑的映射
    format_dir = os.path.join(BASE_DIR, "format")
    try:
        format_files = [f for f in os.listdir(format_dir) if f.lower().endswith('.json')]
        format_map = {os.path.splitext(f)[0].lower(): os.path.join(format_dir, f) for f in format_files}
        print(f"成功載入 {len(format_map)} 個 format JSON 檔案。")
    except FileNotFoundError:
        print(f"錯誤: 找不到 format 目錄: {format_dir}")
        return

    # 2. 初始化 Azure OpenAI 客戶端並讀取系統提示
    client = AzureOpenAI(
        api_key=AZURE_OPENAI_API_KEY,
        api_version=AZURE_OPENAI_API_VERSION,
        azure_endpoint=AZURE_OPENAI_ENDPOINT
    )
    system_prompt_aoai_path = os.path.join(PROMPT_DIR, "prompt_system_using_label.txt")
    system_prompt_aoai = read_prompt_file(system_prompt_aoai_path)

    if not system_prompt_aoai:
        print(f"錯誤：找不到或無法讀取 Azure OpenAI 的系統提示檔案 {system_prompt_aoai_path}，程式終止。")
        return

    # 3. 遍歷 user_input 目錄下的所有 PDF 檔案
    pdf_files = [f for f in os.listdir(USER_INPUT_DIR) if f.lower().endswith(".pdf")]
    if not pdf_files:
        print(f"在 {USER_INPUT_DIR} 中找不到任何 PDF 檔案。")
        return

    print(f"找到 {len(pdf_files)} 個 PDF 檔案，開始進行比對與處理...")

    # 4. 逐一處理每個 PDF 檔案
    for pdf_filename in pdf_files:
        pdf_base_name = os.path.splitext(pdf_filename)[0]
        matched_format_key = None

        # 嘗試在 PDF 檔案名稱中找到匹配的格式鍵
        for format_key in format_map.keys():
            # 使用正則表達式進行更精確的匹配，確保匹配的是整個單詞/短語
            if re.search(r'\b' + re.escape(format_key) + r'\b', pdf_filename, re.IGNORECASE):
                matched_format_key = format_key
                break
        
        # 如果正則表達式沒有找到，則回退到簡單的 'in' 檢查
        if not matched_format_key:
            for format_key in format_map.keys():
                if format_key in pdf_filename.lower():
                    matched_format_key = format_key
                    break

        if matched_format_key:
            print(f"\n- 處理檔案 '{pdf_filename}' (匹配到格式: '{matched_format_key}')")
            json_path = format_map[matched_format_key]
            pdf_path = os.path.join(USER_INPUT_DIR, pdf_filename)

            # 為當前 PDF 建立專用的輸出子目錄
            pdf_output_subdir = os.path.join(OUTPUT_DIR, pdf_base_name)
            os.makedirs(pdf_output_subdir, exist_ok=True)
            print(f"  - 已為檔案 '{pdf_filename}' 建立輸出子目錄: {pdf_output_subdir}")

            doc = None # 初始化 PDF 文件物件
            try:
                # 讀取匹配到的 JSON 設定檔
                with open(json_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)

                doc = fitz.open(pdf_path) # 開啟 PDF 文件
                if not doc.page_count > 0:
                    print("  - 警告: PDF 為空，無法處理。")
                    continue # 跳過此 PDF，處理下一個

                # --- 動作 1: 根據 JSON 設定縮放第一頁圖像 ---
                max_width = config.get('width')
                max_height = config.get('height')
                
                if max_width and max_height:
                    first_page = doc[0]
                    # 以較高 DPI 渲染第一頁以獲得更好的圖像品質
                    pix_first_page = first_page.get_pixmap(dpi=200) 
                    original_image_p1 = Image.frombytes("RGB", [pix_first_page.width, pix_first_page.height], pix_first_page.samples)
                    
                    resized_image = original_image_p1.copy()
                    # 使用 LANCZOS 濾波器進行高品質縮放
                    resized_image.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                    
                    resized_filename = f"{pdf_base_name}_resized.png"
                    resized_path = os.path.join(pdf_output_subdir, resized_filename)
                    resized_image.save(resized_path)
                    print(f"  - 已儲存縮放後的圖片 (第一頁): {resized_filename}")
                else:
                    print("  - 警告: JSON 中缺少 'width' 或 'height' 設定，跳過第一頁縮放。")

                # --- 動作 2: 根據 JSON 設定裁剪指定頁面上的圖像區域 ---
                if 'hints' in config and isinstance(config['hints'], list):
                    for hint in config['hints']:
                        page_num = hint.get('page')
                        field_name = hint.get('field')
                        bbox = hint.get('bbox')

                        # 驗證提示結構是否正確
                        if not (page_num and field_name and bbox and isinstance(page_num, int) and page_num > 0):
                            print(f"  - 警告: 'hints' 中的項目格式不正確或缺少 'page'/'field'/'bbox'。跳過此 hint。")
                            continue
                        
                        # 檢查頁碼是否在 PDF 範圍內
                        if page_num > doc.page_count:
                            print(f"  - 警告: hint 指定的頁面 {page_num} 超出 PDF 總頁數 {doc.page_count}。跳過此 hint。")
                            continue

                        # 獲取目標頁面並轉換為 Pillow 圖像物件
                        target_page = doc[page_num - 1] # PyMuPDF 是 0 索引
                        pix_target_page = target_page.get_pixmap(dpi=200) # 以較高 DPI 渲染
                        image_to_crop = Image.frombytes("RGB", [pix_target_page.width, pix_target_page.height], pix_target_page.samples)

                        # 驗證 bbox 座標格式 (預期為 [x, y, w, h])
                        if not (isinstance(bbox, list) and len(bbox) == 4):
                            print(f"  - 警告: field '{field_name}' 的 bbox 格式不正確，預期為 [x, y, w, h] 陣列。跳過切割。")
                            continue

                        x, y, w, h = bbox[0], bbox[1], bbox[2], bbox[3]
                        
                        # 裁剪框的座標 (左上角 x, 左上角 y, 右下角 x, 右下角 y)
                        crop_box = (x, y, x + w, y + h)
                        
                        # 確保裁剪框在圖像範圍內
                        if crop_box[0] < 0 or crop_box[1] < 0 or crop_box[2] > image_to_crop.width or crop_box[3] > image_to_crop.height:
                            print(f"  - 警告: field '{field_name}' 的 bbox ({x},{y},{w},{h}) 超出頁面 {page_num} 的圖片範圍 ({image_to_crop.width}x{image_to_crop.height})，跳過切割。")
                            continue

                        cropped_image = image_to_crop.crop(crop_box)
                        
                        cropped_filename = f"{field_name}.png"
                        cropped_path = os.path.join(pdf_output_subdir, cropped_filename)
                        cropped_image.save(cropped_path)
                        print(f"  - 已切割並儲存 (頁面 {page_num}, 欄位 '{field_name}'): {cropped_filename}")
                else:
                    print("  - 警告: JSON 中沒有 'hints' 列表或其為空，跳過圖片切割。")
                
                # --- 動作 3: 呼叫 Azure OpenAI 並儲存回應 JSON ---
                print(f"--- 正在為文件 '{pdf_base_name}' 準備 Azure OpenAI 請求 ---")

                # 收集子目錄中所有 PNG 圖像
                image_files = [f for f in os.listdir(pdf_output_subdir) if f.lower().endswith(".png")]
                if not image_files:
                    print(f"  - 在 '{pdf_output_subdir}' 中找不到任何圖片檔案，跳過 Azure OpenAI 請求。")
                    continue

                base64_images_for_aoai = []
                for img_file in image_files:
                    img_path = os.path.join(pdf_output_subdir, img_file)
                    base64_img = image_file_to_base64(img_path)
                    if base64_img:
                        base64_images_for_aoai.append(base64_img)
                
                if not base64_images_for_aoai:
                    print(f"  - 無法編碼 '{pdf_output_subdir}' 中的任何圖片，跳過 Azure OpenAI 請求。")
                    continue

                # 建構用於 OpenAI 呼叫的使用者內容
                user_content_aoai = [{"type": "text", "text": "請根據提供的圖片，提取所有相關資訊，並以 JSON 格式回應。"}]
                user_content_aoai.extend([{"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img}"}} for img in base64_images_for_aoai])

                try:
                    print(f"  - 正在向 Azure OpenAI 發送請求 ({len(base64_images_for_aoai)} 張圖片)... ")
                    response = client.chat.completions.create(
                        model=AZURE_OPENAI_DEPLOYMENT_NAME,
                        messages=[
                            {"role": "system", "content": system_prompt_aoai},
                            {"role": "user", "content": user_content_aoai}
                        ],
                        max_tokens=4096, temperature=0.1, top_p=0.95, response_format={"type": "json_object"}
                    )
                    aoai_json_response = json.loads(response.choices[0].message.content)
                    print("  - 成功收到 Azure OpenAI 回應。")

                    # 儲存 Azure OpenAI 的回應 JSON
                    output_json_filename = f"{pdf_base_name}_with_label.json"
                    output_json_path = os.path.join(pdf_output_subdir, output_json_filename)
                    with open(output_json_path, "w", encoding="utf-8") as f:
                        json.dump(aoai_json_response, f, ensure_ascii=False, indent=4)
                    print(f"  - 已儲存 Azure OpenAI 回應: {output_json_filename}")

                except Exception as e:
                    print(f"  - 呼叫 Azure OpenAI API 或處理回應時發生錯誤: {e}")
                
            finally:
                if doc: # 確保 PDF 文件在開啟後被關閉
                    doc.close()
        else:
            print(f"\n- 檔案 '{pdf_filename}' 未匹配到任何格式，已跳過。")

    print("\n--- 所有檔案處理完畢 ---\n--- 圖片處理任務完成 ---")


if __name__ == "__main__":
    # The original main function for Azure OpenAI processing is preserved.
    # To run it, you can call method_purellm() here.
    method_purellm() 

    # The new function for image processing is called here.
    # method_llm_with_label()
    print("\n--- 圖片處理任務完成 ---")