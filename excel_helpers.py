import os
import json
import re
import shutil
from pathlib import Path
from collections import Counter

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Assuming these are defined in a core shared_helpers or config file
# and will be imported. For now, define them as placeholders.
# In the actual implementation, these will be imported from shared_helpers.py
# from .shared_helpers import EXCEL_OUTPUT_DIR, TOTAL_TEMPLATE_PATH
from shared_helpers import EXCEL_OUTPUT_DIR, TOTAL_TEMPLATE_PATH

def ensure_template_files_exist(log_callback):
    EXCEL_OUTPUT_DIR.mkdir(exist_ok=True)
    if not (EXCEL_OUTPUT_DIR / "total.xlsx").exists():
        try:
            shutil.copy(TOTAL_TEMPLATE_PATH, EXCEL_OUTPUT_DIR / "total.xlsx")
            log_callback(f"[資訊] 已複製總表範本檔案: total.xlsx")
        except FileNotFoundError:
            log_callback(f"[錯誤] 找不到總表範本檔案: {TOTAL_TEMPLATE_PATH}")
            return False
    return True

def sanitize_for_excel(text):
    if not isinstance(text, str): return text
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)

def get_display_value(data_dict):
    if not isinstance(data_dict, dict): return "無"
    if data_dict.get("value"): return data_dict["value"]
    if data_dict.get("derived_value") is not None: return f"{data_dict['derived_value']} (推論)"
    return "無"

def format_evidence(evidence_list):
    if not evidence_list: return ""
    return "\n".join([f"Page {e.get('page', '?')}: \"{e.get('quote', '')}\"" for e in evidence_list])

def format_conflicts(conflicts_list):
    if not conflicts_list: return ""
    return json.dumps(conflicts_list, ensure_ascii=False, indent=2)

def save_total_excel(all_results, log_callback):
    log_callback("--- 開始儲存總表 Excel ---")

    # Helper to find the category for sorting
    def get_category(result_wrapper):
        if not result_wrapper or not result_wrapper.get('processed_data'):
            return None
        for item in result_wrapper['processed_data']:
            # The category is in the 'file' dictionary of the chatgpt_result
            if item.get('chatgpt_result', {}).get('file', {}).get('category'):
                return item['chatgpt_result']['file']['category']
        return None

    # Sort results to have 'Battery Label Artwork' first
    try:
        all_results.sort(key=lambda res: get_category(res) == 'Battery Label Artwork', reverse=True)
        log_callback("  - 已將 'Battery Label Artwork' 資料置頂排序。 ")
    except Exception as e:
        log_callback(f"  - [警告] 排序時發生錯誤: {e}")

    try:
        total_output_path = EXCEL_OUTPUT_DIR / "total.xlsx"
        total_wb = load_workbook(total_output_path)
        total_ws = total_wb.active
        for result_wrapper in all_results:
            if not result_wrapper or 'processed_data' not in result_wrapper or not result_wrapper['processed_data']:
                continue
            
            # Merge results from all pages/hints of a single PDF
            merged_data = {}
            for item_result in result_wrapper['processed_data']:
                # The actual data is nested in 'chatgpt_result'
                if 'chatgpt_result' in item_result:
                    merged_data.update(item_result['chatgpt_result'])

            if not merged_data: continue

            # Add file-level info
            if 'file' not in merged_data: merged_data['file'] = {}
            merged_data['file']['name'] = result_wrapper['file_name']
            data = merged_data

            row_data = [
                sanitize_for_excel(data.get('file', {}).get('name', '')),
                sanitize_for_excel(data.get('file', {}).get('category', '')),
                sanitize_for_excel(data.get('model_name', {}).get('value', '')),
                sanitize_for_excel(get_display_value(data.get('nominal_voltage_v', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_batt_capacity_wh', {}))),
                sanitize_for_excel(get_display_value(data.get('typ_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_capacity_mah', {}))),
                sanitize_for_excel(get_display_value(data.get('rated_energy_wh', {}))),
                sanitize_for_excel(data.get('notes', '')),
                sanitize_for_excel(format_conflicts(data.get('conflicts', [])))
            ]
            total_ws.append(row_data)

        total_wb.save(total_output_path)
        log_callback(f"  - 總表 {total_output_path.name} 已儲存")
        return total_output_path
    except Exception as e:
        log_callback(f"[錯誤] 儲存總表 Excel 檔案時發生錯誤: {e}")
        return None

def apply_highlighting_rules(excel_path, log_callback):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        log_callback(f"讀取 Excel 檔案成功: {os.path.basename(excel_path)}")
    except FileNotFoundError:
        log_callback(f"[錯誤] 找不到要進行標色的 Excel 檔案: {excel_path}")
        return

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    category_col_idx = 2
    columns_to_check_indices = [3, 4, 5, 6, 7, 8]

    artwork_row_indices = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        if len(row) >= category_col_idx and row[category_col_idx - 1] == 'Battery Label Artwork':
            artwork_row_indices.append(row_idx)

    if ws.max_row <= 1:
        log_callback("[資訊] Excel 中沒有資料, 跳過標色 ")
        return

    if len(artwork_row_indices) == 1:
        log_callback("[規則 1] 偵測到單一 'Battery Label Artwork', 以此為標準 ")
        standard_row_idx = artwork_row_indices[0]
        for col_idx in columns_to_check_indices:
            standard_cell = ws.cell(row=standard_row_idx, column=col_idx)
            standard_value = standard_cell.value
            for row_idx in range(2, ws.max_row + 1):
                if row_idx == standard_row_idx: continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value != standard_value:
                    cell.fill = red_fill

    elif len(artwork_row_indices) > 1:
        log_callback("[規則 2] 偵測到多筆 'Battery Label Artwork', 進行內部比對 ")
        for col_idx in columns_to_check_indices:
            artwork_values = [ws.cell(row=r_idx, column=col_idx).value for r_idx in artwork_row_indices]
            if len(set(artwork_values)) > 1:
                header_name = ws.cell(row=1, column=col_idx).value
                log_callback(f"  - 欄位 '{header_name}' 在 Artwork 中發現不一致, 全部標紅 ")
                for r_idx in artwork_row_indices:
                    ws.cell(row=r_idx, column=col_idx).fill = red_fill

    else: # No artwork rows
        log_callback("[規則 3] 未偵測到 'Battery Label Artwork' , 採用多數決 ")
        for col_idx in columns_to_check_indices:
            col_values = [ws.cell(row=r_idx, column=col_idx).value for r_idx in range(2, ws.max_row + 1) if ws.cell(row=r_idx, column=col_idx).value is not None]
            if not col_values: continue
            value_counts = Counter(col_values)
            most_common = value_counts.most_common()
            header_name = ws.cell(row=1, column=col_idx).value
            if len(most_common) > 1 and most_common[0][1] == most_common[1][1]:
                log_callback(f"  - 欄位 '{header_name}' 出現平手, 整欄標紅 ")
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=col_idx).fill = red_fill
            else:
                majority_value = most_common[0][0]
                log_callback(f"  - 欄位 '{header_name}' 的多數值為 '{majority_value}' ")
                for r_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    if cell.value != majority_value:
                        cell.fill = red_fill

    wb.save(excel_path)
    log_callback("成功儲存已標色的 Excel 檔案 ")
